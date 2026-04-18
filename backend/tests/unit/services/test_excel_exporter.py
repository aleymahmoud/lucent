"""Tests for Excel export — ensure files open as valid xlsx with expected sheets."""
from __future__ import annotations

from datetime import datetime

import openpyxl
import pytest

from app.schemas.forecast import (
    CrossValidationResultResponse,
    ForecastMethod,
    ForecastResultResponse,
    ForecastStatus,
    MetricsResponse,
    ModelSummaryResponse,
    PredictionResponse,
    BatchForecastStatusResponse,
)
from app.services.excel_exporter import export_batch, export_single


def _build_result(entity_id: str = "ENTITY_A", with_cv: bool = False) -> ForecastResultResponse:
    predictions = [
        PredictionResponse(
            date=f"2024-01-{i + 1:02d}",
            value=100.0 + i,
            lower_bound=95.0 + i,
            upper_bound=105.0 + i,
        )
        for i in range(5)
    ]
    metrics = MetricsResponse(mae=1.2, rmse=1.5, mape=2.3, mse=2.25, r2=0.9)
    model_summary = ModelSummaryResponse(
        method="arima",
        parameters={"p": 1, "d": 1, "q": 1},
        coefficients=[
            {"name": "ar.L1", "estimate": 0.45, "std_error": 0.08, "p_value": 0.01, "significant": True}
        ],
        diagnostics={"residual_mean": 0.01, "residual_std": 1.0},
        regressors_used=None,
        residuals=[0.1, -0.2, 0.15, -0.05, 0.12],
    )
    cv_results = None
    if with_cv:
        fold_metrics = MetricsResponse(mae=1.3, rmse=1.6, mape=2.5)
        cv_results = CrossValidationResultResponse(
            folds=2,
            method="rolling",
            metrics_per_fold=[fold_metrics, fold_metrics],
            average_metrics=fold_metrics,
        )
    return ForecastResultResponse(
        id="f-1",
        dataset_id="d-1",
        entity_id=entity_id,
        method=ForecastMethod.ARIMA,
        status=ForecastStatus.COMPLETED,
        progress=100,
        predictions=predictions,
        metrics=metrics,
        model_summary=model_summary,
        cv_results=cv_results,
        created_at=datetime.utcnow(),
    )


def test_export_single_produces_valid_xlsx():
    result = _build_result(with_cv=False)
    buf = export_single(result)
    wb = openpyxl.load_workbook(buf)

    # Expected sheets (no CV sheet when CV disabled)
    assert "Forecast" in wb.sheetnames
    assert "Metrics" in wb.sheetnames
    assert "Model" in wb.sheetnames
    assert "Cross-Validation" not in wb.sheetnames

    # Forecast sheet has header + 5 data rows
    ws = wb["Forecast"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Date", "Value", "Lower", "Upper")
    assert len(rows) == 6


def test_export_single_with_cv_has_cv_sheet():
    result = _build_result(with_cv=True)
    buf = export_single(result)
    wb = openpyxl.load_workbook(buf)
    assert "Cross-Validation" in wb.sheetnames
    ws = wb["Cross-Validation"]
    rows = list(ws.iter_rows(values_only=True))
    # header + 2 folds + 1 average
    assert len(rows) >= 4


def test_export_batch_includes_summary_all_data_and_per_entity_sheets():
    r1 = _build_result(entity_id="ALPHA")
    r2 = _build_result(entity_id="BETA")
    batch = BatchForecastStatusResponse(
        batch_id="b-1",
        total=2,
        completed=2,
        failed=0,
        in_progress=0,
        status=ForecastStatus.COMPLETED,
        results=[r1, r2],
    )
    buf = export_batch(batch)
    wb = openpyxl.load_workbook(buf)
    names = set(wb.sheetnames)

    assert "Summary" in names
    assert "All_Data" in names
    # Per-entity sheets (names may be truncated/deduplicated but must exist)
    assert any(s.startswith("ALPHA") for s in names)
    assert any(s.startswith("BETA") for s in names)

    # Summary has one row per entity + header
    rows = list(wb["Summary"].iter_rows(values_only=True))
    assert len(rows) >= 3  # header + 2 entities


def test_export_batch_handles_long_and_duplicate_entity_names():
    # Excel caps sheet names at 31 chars — exporter must dedupe
    long = "X" * 40
    batch = BatchForecastStatusResponse(
        batch_id="b-2",
        total=2,
        completed=2,
        failed=0,
        in_progress=0,
        status=ForecastStatus.COMPLETED,
        results=[_build_result(entity_id=long), _build_result(entity_id=long)],
    )
    buf = export_batch(batch)
    wb = openpyxl.load_workbook(buf)
    # All sheet names must be unique + <= 31 chars
    assert len(wb.sheetnames) == len(set(wb.sheetnames))
    for n in wb.sheetnames:
        assert len(n) <= 31
